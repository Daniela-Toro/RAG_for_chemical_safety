# functions.py
# Necessary imports
# Standard Library
import os
import re
import json
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from config import folder_documents, JSON_PATHS, template_path, output_Excel
from llm_setup import llm as default_llm, llm, db
from utils import (
    _FIELD_PATTERNS,
    PPE_FIELDS,
    _STORAGE_PATTERNS,
    STORAGE_FIELDS,
    dtr_tables, hazards_protection_measures_fields, hazards_fields_dtr,
    waste_disposal_measures_fields_dtr, spill_management_fields_dtr, fire_procedures_fields_dtr,
    first_aid_procedures_fields_dtr, storage_fields_dtr, hazards_fields_statements,
)
def process_document(source_match, content):
    """
    Processes an SDS/MSDS document from data extraction to Excel completion.
    Args:
        source_match (str): File name or identifier.
        content (str): Full content of the document.
    Returns:
        tuple: (dict of updated JSONs, bool indicating if Excel was generated)
    """

    # Read JSONs
    json_data = {}
    try:
        for key, path in JSON_PATHS.items():
            with open(path, "r", encoding="utf-8") as f:
                json_data[key] = json.load(f)
    except Exception as e:
        raise RuntimeError(f"Error reading base JSONs: {e}")

    # Later access to each JSON:
    json_hazards = json_data["hazards"]
    json_waste_disposal_measures = json_data["waste_disposal_measures"]
    json_spill_management = json_data["spill_management"]
    json_fire_procedures = json_data["fire_procedures"]
    json_first_aid_procedures = json_data["first_aid_procedures"]
    json_storage = json_data["storage"]

    # 1. Document identification
    print("Document identification")
    base_id = get_document_id(source_match)

    # 2. Initial information extraction
    print("Initial information extraction")
    chemical_names = extract_chemical_names(source_match, content)

    # 3. Initialization of JSONs with base data
    print("Initialization of JSONs with base data")
    updated_json_hazards = fill_json_chemical_fields(
        json_input=json_hazards,
        content=content,
        base_id=base_id,
        chemical_names=chemical_names,
        source_match=source_match
    )

    updated_json_waste_disposal_measures = fill_json_chemical_fields(
        json_input=json_waste_disposal_measures,
        content=content,
        base_id=base_id,
        chemical_names=chemical_names,
        source_match=source_match
    )

    updated_json_spill_management = fill_json_chemical_fields(
        json_input=json_spill_management,
        content=content,
        base_id=base_id,
        chemical_names=chemical_names,
        source_match=source_match
    )

    updated_json_fire_procedures = fill_json_chemical_fields(
        json_input=json_fire_procedures,
        content=content,
        base_id=base_id,
        chemical_names=chemical_names,
        source_match=source_match
    )

    updated_json_first_aid_procedures = fill_json_chemical_fields(
        json_input=json_first_aid_procedures,
        content=content,
        base_id=base_id,
        chemical_names=chemical_names,
        source_match=source_match
    )

    updated_json_storage = fill_json_chemical_fields(
        json_input=json_storage,
        content=content,
        base_id=base_id,
        chemical_names=chemical_names,
        source_match=source_match
    )

    # 4. Processing fields with images / measures
    print("Processing fields with images / measures")
    updated_json_hazards = control_measures_with_images(
        "Personal Protection",
        content,
        hazards_protection_measures_fields,
        updated_json_hazards['Sheet_2'],
        llm
    )

    fields_with_images(
        field_name="Hazard Statements",
        content=content,
        fields_list=hazards_fields_statements,
        data_dict=updated_json_hazards['Sheet_2'],
        model=llm
    )

    updated_json_storage = storage_fields_with_images(
        "Storage",
        content,
        STORAGE_FIELDS,
        updated_json_storage['Sheet_2'],
        llm
    )

    # 5. Enrichment with Hazard Group RAG
    print("Enrichment with Hazard Group RAG")
    updated_json_hazards = fill_hazard_group_rag(source_match, updated_json_hazards, content)
    updated_json_waste_disposal_measures = fill_hazard_group_rag(source_match, updated_json_waste_disposal_measures, content)
    updated_json_storage = fill_hazard_group_rag(source_match, updated_json_storage, content)

    # 6. Filling severity / probability fields
    print("Filling severity / probability fields")
    fill_json_severity_probability(updated_json_hazards)

    # 7. Extraction of specific text by section
    print("Extraction of specific text by section")
    updated_json_hazards = extract_hazards_text(
        source_match,
        updated_json_hazards,
        model=llm,
        content=content,
        fields_list=hazards_fields_dtr
    )

    updated_json_waste_disposal_measures = general_text_extraction(
        source_match,
        updated_json_waste_disposal_measures,
        model=llm,
        content=content,
        fields_list=waste_disposal_measures_fields_dtr,
        table_index=1
    )

    updated_json_spill_management = general_text_extraction(
        source_match,
        updated_json_spill_management,
        model=llm,
        content=content,
        fields_list=spill_management_fields_dtr,
        table_index=2
    )

    updated_json_fire_procedures = general_text_extraction(
        source_match,
        updated_json_fire_procedures,
        model=llm,
        content=content,
        fields_list=fire_procedures_fields_dtr,
        table_index=3
    )

    updated_json_first_aid_procedures = general_text_extraction(
        source_match,
        updated_json_first_aid_procedures,
        model=llm,
        content=content,
        fields_list=first_aid_procedures_fields_dtr,
        table_index=4
    )

    updated_json_storage = general_text_extraction(
        source_match,
        updated_json_storage,
        model=llm,
        content=content,
        fields_list=storage_fields_dtr,
        table_index=5
    )

    # 8. Prepare list of JSONs for Excel
    print("Prepare list of JSONs for Excel")
    list_of_jsons_to_excel = [
        updated_json_waste_disposal_measures,
        updated_json_storage,
        updated_json_fire_procedures,
        updated_json_first_aid_procedures,
        updated_json_hazards,
        updated_json_spill_management
    ]

    print(updated_json_hazards)
    print(updated_json_waste_disposal_measures)
    print(updated_json_spill_management)
    print(updated_json_fire_procedures)
    print(updated_json_first_aid_procedures)
    print(updated_json_storage)

    # 9. Create / fill final Excel
    print("Create / fill final Excel")
    excel_created = fill_excel_with_json(
        list_of_jsons_to_excel,
        template_path,
        output_Excel,
        source_match=source_match
    )

    # 10. Return updated JSONs and Excel status
    updated_jsons = {
        "Hazards": updated_json_hazards,
        "Waste_disposal_measures": updated_json_waste_disposal_measures,
        "Storage": updated_json_storage,
        "Fire_procedures": updated_json_fire_procedures,
        "First_aid_procedures": updated_json_first_aid_procedures,
        "Spill_management": updated_json_spill_management
    }

    return updated_jsons, excel_created
# Filtering
def list_db_sources(db):
    """
    Lists all unique 'source' entries present in a Chroma database.
    Args:
        db: Chroma database instance.
    Returns:
        A set of unique source strings extracted from the database metadata.
    Notes:
        - Prints the total number of unique sources found.
        - Prints up to the first 100 sources for quick inspection.
    """

    res = db.get()
    metadatas = res.get("metadatas", [])
    sources = {meta.get("source", "") for meta in metadatas if isinstance(meta, dict)}
    print(f"Found {len(sources)} unique sources")
    for s in list(sources)[:100]:
        print(" -", repr(s))
    return sources

# Function responsible for returning a document for the retriever (with content)
def filter_document(query_doc, db, k=10):
    """
    Retrieves the most relevant document from a vector database based on a query,
    and returns its content along with the source filename.
    Args:
        query_doc (str): The query text used to find similar documents.
        db: Vector database object with an `as_retriever` method.
        k (int, optional): Number of similar documents to fetch. Defaults to 10.
    Returns:
        tuple: (source_match, content)
            - source_match (str): The filename of the most similar document.
            - content (str or None): The full text content of the document, or None if
              the file cannot be found or read.
    Raises:
        ValueError: If no similar documents are found in the database.

    Notes:
        - Uses the database retriever with similarity search.
        - Assumes that the file exists in `folder_documents` with the name in metadata.
        - If the physical file is missing or cannot be read, returns None for content.
    """
    print(db)
    retriever = db.as_retriever(search_type="similarity", search_kwargs={"k": k})
    relevant_documents = retriever.invoke(query_doc)

    if not relevant_documents:
        raise ValueError("No similar document was found in the database.")

    # Full name of the document (according to metadata)
    source_match = relevant_documents[0].metadata.get('source')
    print(f"Most similar document: {source_match}")

    # Physical path to the file
    doc_path = os.path.join(folder_documents, source_match)

    if not os.path.exists(doc_path):
        print(f"Physical file not found: {doc_path}")
        return source_match, None

    try:
        with open(doc_path, "r", encoding="utf-8") as file:
            content = file.read()
        return source_match, content
    except Exception as e:
        print(f"Error reading the document: {e}")
        return source_match, None


# Chemical Name and SDS
# Function to get the document ID
def get_document_id(source_match: str) -> str:
    """
    Generates a base document ID from the given source string.
    Args:
        source_match (str): The source string (e.g., filename or document identifier).
    Returns:
        str: The first 14 characters of the source string, used as a base document ID.
    Notes:
        - Prints the extracted base ID for verification.
    """
    # Extract the first 14 characters
    base_id = source_match[:14]
    print(f"Extracted base ID: {base_id}")
    return base_id

# Function to get the product name
def get_product_name(source_match: str) -> str:
    """
    Extracts the product name from a given source string, typically a file name.
    Args:
        source_match (str): The source string or file name.
    Returns:
        str: The extracted product name, ignoring the first 15 characters and the last 4
             characters (typically the file extension).
    Notes:
        - Prints the extracted product name for verification.
        - Assumes the file name is at least 19 characters long (otherwise slicing may produce unexpected results).
    """

    # Slice from character 15 to 3 characters from the end
    product_name = source_match[15:-3]
    print(f"Product name: {product_name}")
    return product_name

# Clean raw chemical names to remove noise
def clean_chemical_names(raw_list):
    """
    Cleans a list of raw chemical names by removing noise, irrelevant entries, and duplicates.
    Args:
        raw_list (list of str): Raw chemical names extracted from documents or databases.
    Returns:
        list of str: A list of cleaned chemical names.
    Cleaning rules:
        - Trims leading and trailing spaces, capitalizes words.
        - Ignores names shorter than 4 or longer than 60 characters.
        - Filters out names containing common noise words like 'No Substance', 'Regulation', 'Explosives', etc.
        - Removes leading punctuation or asterisks.
        - Collapses multiple spaces into a single space.
        - Keeps only alphanumeric characters, spaces, hyphens, parentheses, and commas.
        - Removes duplicates while preserving order.
    """
    clean_names = []
    noisy_words = [
        "No Substance", "Regulation", "Annex", "List", "Assessed", "Authorisation",
        "Candidate", "Pop", "Pic", "Explosives", "Drug", "Ozone", "###"
    ]
    for name in raw_list:
        name = name.strip().title()
        if len(name) < 4 or len(name) > 60:
            continue
        if any(word.lower() in name.lower() for word in noisy_words):
            continue
        name = re.sub(r"^\*\*[:\-]?\s*", "", name)
        name = re.sub(r"\s{2,}", " ", name)
        if not re.match(r'^[A-Za-z0-9\s\-\(\),]+$', name):
            continue
        if name not in clean_names:
            clean_names.append(name)
    return clean_names

# Extract chemical names from a document (SDS/MSDS)
def extract_chemical_names(source_match, content, use_llm=True, model=None):
    """
    Extracts chemical ingredient names from an SDS/MSDS document using a combination of regex and LLM-based parsing.
    Args:
        source_match (str): File name or identifier of the SDS/MSDS document.
        content (str): Full text content of the SDS/MSDS document.
        use_llm (bool): If True, uses an LLM to supplement extraction when regex finds few or no names.
        model: Instance of an LLM (e.g., ChatOpenAI) to use if LLM extraction is enabled.
               If None and use_llm=True, uses the global default llm instance.
    Returns:
        list of str: Cleaned list of extracted chemical names.
    """
    if model is None and use_llm:
        model = default_llm

    found_names = []

    # Step 1: Regex on Section 3
    section3 = ""
    match = re.search(r"(section\s*3.*?composition.*?)(section\s*\d+|$)", content, re.IGNORECASE | re.DOTALL)
    if match:
        section3 = match.group(1)
        patterns = [
            r"ingredient[s]?:?\s*([\w\s\-\(\)\/]+)",
            r"component[s]?:?\s*([\w\s\-\(\)\/]+)",
            r"substance\s*name:?\s*([\w\s\-\(\)\/]+)"
        ]
        for pattern in patterns:
            matches = re.findall(pattern, section3, re.IGNORECASE)
            for name in matches:
                name = name.strip().title()
                if name and name not in found_names:
                    found_names.append(name)

    # Step 2: LLM if regex is insufficient
    if use_llm and (not found_names or len(found_names) < 2):
        prompt_template = f"""
        You are an expert assistant for analyzing chemical safety datasheets (SDS / MSDS).

        Task: Extract **only the chemical ingredient names** from the following document.

        Strict rules:
        - Focus on Section 3 (Composition / Information on Ingredients)
        - Ignore the general Product Name (e.g., WD-40 Multi-Use Product Aerosol)
        - Extract only **individual chemical names**
        - DO NOT include percentages, CAS numbers, regulatory phrases, or comments
        - ALWAYS return valid JSON in this schema:

        {{ "chemical_names": ["Name1", "Name2", "Name3"] }}

        Document:
        ---
        {content}
        ---
        """
        response = model.invoke(prompt_template).content

        # Clean ```json ... ```
        clean_response = response.strip().strip("`").replace("json", "").strip()

        try:
            data = json.loads(clean_response)
            llm_names = data.get("chemical_names", [])
            for name in llm_names:
                name = name.strip().title()
                if name and name not in found_names:
                    found_names.append(name)
        except Exception as e:
            print(f"Could not parse LLM response: {e}")
            print("Raw LLM response:", response)

    # Step 3: Final cleaning
    final_names = []
    noisy_words_final = ["Not Hazardous", "No Substance", "See Section", "###", "Ltd", "Com"]
    for name in found_names:
        if len(name) < 3 or len(name) > 80:
            continue
        if any(word.lower() in name.lower() for word in noisy_words_final):
            continue
        name = re.sub(r"\s{2,}", " ", name).strip()
        if name not in final_names:
            final_names.append(name)

    print(f"Document used: {source_match}")
    print(f"Final chemical names: {final_names}")
    return final_names


# Function to combine the product name with its list of chemical compounds
def extract_product_info(source_match: str, content: str) -> dict:
    """
    Extracts and combines the product name with its chemical ingredients from an SDS/MSDS document.
    Args:
        source_match (str): File name or identifier of the SDS/MSDS document.
        content (str): Full text content of the SDS/MSDS document.
    Returns:
        dict: Dictionary containing:
            - "product_name": the cleaned product name extracted from the file name.
            - "chemical_names": a list of chemical compounds found in Section 3 of the SDS/MSDS.
    Procedure:
        1. Extract the product name from the file name, ignoring irrelevant prefixes or extensions.
        2. Extract chemical names using `extract_chemical_names`, which uses regex and optionally an LLM.
        3. Combine both pieces of information into a single dictionary for further processing or storage.
    """

    # Get the base product name
    product_name = get_product_name(source_match)

    # Get chemical compounds
    chemical_names = extract_chemical_names(source_match, content)

    # Build dictionary
    result = {
        "product_name": product_name,
        "chemical_names": chemical_names
    }

    return result

def fill_json_chemical_fields(
    json_input: Dict[str, Any],
    content: str,
    base_id: str,
    chemical_names: Optional[List[str]] = None,
    source_match: Optional[str] = None
) -> Dict[str, Any]:
    """
    Searches and fills the 'chemical_name' and 'sds_reference' fields
    within a nested JSON structure without modifying other fields.

    Args:
        json_input (Dict[str, Any]): The nested JSON structure to update.
        content (str): Full text content of the SDS/MSDS document.
        base_id (str): Unique identifier extracted from the document or file name.
        chemical_names (Optional[List[str]]): List of chemical names to fill.
            If not provided, they are extracted from the document.
        source_match (Optional[str]): File name or identifier of the SDS/MSDS document,
            used to extract the product name.

    Returns:
        Dict[str, Any]: The same JSON structure with updated 'chemical_name'
        and 'sds_reference' fields.

    Procedure:
        1. If `chemical_names` is not provided, extract them from the document using `extract_chemical_names`.
        2. Extract the product name from `source_match` if available.
        3. Recursively traverse the JSON structure:
            - For each 'chemical_name' field, fill it with either:
                "{product_name}: name1, name2, ..." or "name1, name2, ..." if product name is unavailable.
            - For each 'sds_reference' field, fill it with `base_id`.
        4. Preserve the original structure of the JSON for all other fields.
    """

    if chemical_names is None:
        chemical_names = extract_chemical_names(source_match, content)

    product_name = get_product_name(source_match) if source_match else None

    def update_fields(data):
        if isinstance(data, dict):
            for key, value in data.items():

                # 🔹 Store as plain string
                if key == "chemical_name" and isinstance(value, dict):
                    if product_name:
                        chemicals_str = f"{product_name}: {', '.join(chemical_names)}"
                        value["response"] = chemicals_str
                        value["to_excel"] = chemicals_str
                    else:
                        chemicals_str = ", ".join(chemical_names)
                        value["response"] = chemicals_str
                        value["to_excel"] = chemicals_str

                elif key == "sds_reference" and isinstance(value, dict):
                    value["response"] = base_id
                    value["to_excel"] = base_id

                else:
                    update_fields(value)

        elif isinstance(data, list):
            for item in data:
                update_fields(item)

    update_fields(json_input)
    return json_input


# Control Measures
def _candidate_lines(text: str):
    """
    Extracts useful lines from a text document by cleaning and normalizing each line.
    Args:
        text (str): Raw text input, possibly containing bullets, leading/trailing whitespace,
                    or other list symbols.
    Returns:
        list: A list of cleaned lines, stripped of leading/trailing whitespace and common bullet
              or list markers, ready for further processing.
    Procedure:
        1. Split the text into lines.
        2. Remove empty lines.
        3. Strip leading/trailing whitespace from each line.
        4. Remove common list markers or bullets such as '-', '*', '•', etc.
        5. Return the resulting list of cleaned lines.
    """

    lines = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        line = re.sub(r"^[\-\*\•\u2022]+\s*", "", line).strip()
        lines.append(line)
    return lines

def _matches_any(line: str, patterns):
    """
    Checks whether a given text line matches any of the provided regular expression patterns.
    Args:
        line (str): The text line to check.
        patterns (list): A list of regex patterns (strings) to match against the line.
    Returns:
        bool: True if at least one pattern matches the line, False otherwise.
    Procedure:
        1. Convert the line to lowercase for case-insensitive matching.
        2. Iterate over all patterns.
        3. Use `re.search` to check for a match with each pattern.
        4. Return True on the first match found; otherwise, return False.
    """

    line_lower = line.lower()
    return any(re.search(p, line_lower) for p in patterns)

def _find_support_for_field(summary_text: str, field_key: str) -> str:
    """
    Finds the first line in the provided text that matches the patterns associated with a given field.
    Args:
        summary_text (str): The text block or summary from which to extract supporting lines.
        field_key (str): The key of the field for which we want supporting evidence.
    Returns:
        str: The first matching line supporting the field. If no match is found, returns an empty string.
    Procedure:
        1. Retrieve the regex patterns associated with `field_key` from `_FIELD_PATTERNS`.
        2. If no patterns exist for this field, return an empty string.
        3. Split the text into candidate lines using `candidate_lines`.
        4. Check each line against all patterns using `matches_any`.
        5. Return the first line that matches; if none match, return "".
    """

    patterns = _FIELD_PATTERNS.get(field_key, [])
    if not patterns:
        return ""
    for line in candidate_lines(summary_text):
        if matches_any(line, patterns):
            return line.strip()
    return ""

def _extract_json_block(text: str):
    """
    Extracts a JSON block from a given text string.
    Args:
        text (str): The input text potentially containing JSON data.
                    Supports JSON wrapped in ```json ... ``` or plain JSON.
    Returns:
        str: The JSON string extracted from the text. If no explicit JSON block is found,
             returns the original text as a fallback.
    Procedure:
        1. Attempt to locate a JSON block using a regex that captures from '{' to '}' at the end of the text.
        2. If a match is found, return the matched JSON string.
        3. If no match is found, return the original text as a fallback for parsing.
    """

    match = re.search(r"\{.*\}\s*$", text, re.DOTALL)
    if match:
        return match.group(0)
    # Plan B: try parsing as is
    return text

def control_measures_with_images(field_name, content, fields_list, data_dict, use_llm=True, model=None):
    """
    Extracts control measures (including PPE) from a text context and fills the JSON accordingly.
    This function accepts either:
      - the full JSON that contains a "Sheet_2" key, or
      - the inner sheet dict itself (i.e. what would be json['Sheet_2']).
    It always returns a full JSON with "Sheet_2" as top-level key, so callers that do:
        updated_json = control_measures_with_images(..., updated_json['Sheet_2'], ...)
    will NOT lose the Sheet_2 wrapper.
    """
    # Decide si nos pasaron el JSON completo o solo el sheet
    wrapped_input = True
    if isinstance(data_dict, dict) and "Sheet_2" in data_dict:
        sheet = data_dict["Sheet_2"]
    else:
        wrapped_input = False
        # trabajar sobre una copia superficial para evitar efectos colaterales indeseados
        sheet = dict(data_dict) if isinstance(data_dict, dict) else {}

    # Asegurar que sheet es un dict válido
    if sheet is None or not isinstance(sheet, dict):
        sheet = {}

    # Preparar el modelo LLM
    if model is None and use_llm:
        model = default_llm

    # Base summary
    base_prompt = """
    Answer STRICTLY using only the content retrieved from the provided context.
    Do not invent or add external information.
    If the context contains no information relevant to the question, state explicitly that the information is not available.
    """
    request = f"""
    Answer the question based only on these instructions: {base_prompt}.
    What are the main {field_name} risks or measures in the context: {content}?
    Answer in bullet points, keeping the exact wording from the context whenever possible.
    """
    field_summary = model.invoke(request).content

    # Special case: Hazard Statements (trabajar sobre sheet)
    if field_name == 'Hazard Statements' and "hazard_statements" in sheet:
        sheet["hazard_statements"]["to_excel"] = field_summary.replace('*','')

    # Prompt PPE: binary mapping of 6 fields
    mapping_prompt = f"""
    Based only on this extracted information:
    {field_summary}

    Check which of the following protection measures are explicitly required or implied.
    Mark with 'X' if true, otherwise '' (empty string).

    Fields:
    - wear_full_face_visor: full face visor, face shield
    - box_goggles_must_be_worn: eye protection, goggles, safety glasses
    - protective_gloves_must_be_worn: protective gloves, hand protection
    - laboratory_coats_must_be_worn: lab coat, protective clothing, body protection
    - use_local_exhaust_ventilation: local exhaust ventilation, fume hood
    - no_open_flames: no open flames, keep away from ignition sources

    Respond EXACTLY with lines like:
    field_name: X
    field_name:
    (one per line; no extra commentary)
    """
    ppe_result = model.invoke(mapping_prompt).content

    # Initialize PPE fields empty in sheet (si no existen, crear estructura mínima)
    for field in fields_list:
        if field not in sheet or not isinstance(sheet[field], dict):
            sheet[field] = {"content": "", "position": "", "response": "", "to_excel": ""}
        else:
            sheet[field]["to_excel"] = ""
            sheet[field]["response"] = ""

    # Mark PPE fields and add evidence
    for line in ppe_result.splitlines():
        if ":" not in line:
            continue
        field, value = line.split(":", 1)
        field = field.strip().lstrip("-").strip()
        value = value.strip()
        if field in fields_list and field != "other_control_measures":
            if value == "X":
                sheet[field]["to_excel"] = "X"
                sheet[field]["response"] = _find_support_for_field(field_summary, field) or ""

    # Prompt "Other control measures" (excluding the six PPE fields)
    active_ppe = [f for f in PPE_FIELDS if sheet.get(f, {}).get("to_excel") == "X"]
    other_prompt = f"""
    You are given this extracted text (context):
    {field_summary}

    Task: List ALL explicit control or prevention measures that are NOT any of these categories:
    - full face visor / face shield
    - eye protection / goggles / safety glasses
    - protective gloves / hand protection
    - lab coat / protective clothing / body protection
    - local exhaust ventilation / fume hood
    - no open flames / ignition sources

    Also, avoid simply rephrasing measures already covered by these categories, even if they are active: {active_ppe}.

    Return STRICT JSON with this schema:
    {{
      "list": ["short, practical measure 1", "short, practical measure 2", ...],
      "paragraph": "exact paragraph(s) from the context where those 'other' measures appear"
    }}

    Requirements:
    - "list": concise, actionable phrases; do not include H-codes or P-codes; no headings like 'Hygiene:'.
    - "paragraph": copy the original text fragment(s) verbatim from the context containing those 'other' measures.
    - If there are none, return: {{ "list": [], "paragraph": "" }}.
    """
    other_raw = model.invoke(other_prompt).content

    # Robust JSON parsing
    other_json = {"list": [], "paragraph": ""}
    try:
        block = extract_json_block(other_raw)
        other_json = json.loads(block)
        if not isinstance(other_json, dict):
            other_json = {"list": [], "paragraph": ""}
    except Exception:
        other_json = {"list": [], "paragraph": ""}

    # Normalize types
    measures_list = other_json.get("list", [])
    if not isinstance(measures_list, list):
        measures_list = []
    paragraph = other_json.get("paragraph", "")
    if not isinstance(paragraph, str):
        paragraph = ""

    # Write other_control_measures into sheet (si existe)
    if "other_control_measures" in sheet:
        sheet["other_control_measures"]["to_excel"] = "; ".join(m.strip() for m in measures_list).strip(" ;")
        sheet["other_control_measures"]["response"] = paragraph if paragraph.strip() else "; ".join(measures_list).strip(" ;")
    else:
        # If it does not exist, create it to maintain consistency
        sheet["other_control_measures"] = {
            "content": "Other control measures (inherent or additionally required)",
            "position": "",
            "response": paragraph if paragraph.strip() else "; ".join(measures_list).strip(" ;"),
            "to_excel": "; ".join(m.strip() for m in measures_list).strip(" ;")
        }

    # If the original input was the full JSON, update its Sheet_2 and return the full JSON
    if wrapped_input:
        data_dict["Sheet_2"] = sheet
        return data_dict

    # If the input was only the sheet, return a JSON containing Sheet_2 to avoid breaking assignments
    return {"Sheet_2": sheet}

# Hazards Extraction
def fields_with_images(field_name, content, fields_list, data_dict, use_llm=True, model=None):
    """
    Updates hazard/pictogram fields in a JSON based on LLM analysis.
    Args:
        field_name (str): Name of the main field being processed (e.g., 'Hazard Statements').
        content (str): Text context from which hazards and measures are extracted.
        fields_list (list): List of keys in `data_dict` corresponding to hazard/pictogram fields.
        data_dict (dict): JSON structure where the results will be stored.
                          Each key must contain 'content', 'position', and 'to_excel'.
        llm: Language model object with an 'invoke' or 'predict' method to get LLM responses.
    Behavior:
        1. Generates a base response from the LLM describing the main risks/measures in the context.
        2. If `field_name` is 'Hazard Statements', updates the `hazard_statements` entry in `to_excel`.
        3. Iterates over each field in `fields_list` (pictograms/hazard indicators):
            - Prompts the LLM to check if the risk is explicitly mentioned.
            - Marks 'X' in `to_excel` if present, otherwise leaves it empty.
        4. Updates the original `data_dict` with the new 'to_excel' values for all fields.
    Returns:
        None: Updates `data_dict` in place.
    """
    if model is None and use_llm:
        model = default_llm

    update_dict = {}

    base_prompt = f"""
    Answer STRICTLY using only the content retrieved from the provided context.
    Do not invent or add external information.
    If the context contains no information relevant to the question, state explicitly that the information is not available.
    """
    request = f"""
    Answer the question based only on these instructions: {base_prompt}.
    What are the main {field_name} risks or measures in the context: {content}?
    Answer only the explicit values and exclude other precautions.
    """

    # Get base response from LLM
    field_response = model.invoke(request).content

    # Save hazard statements text if applicable
    if field_name == 'Hazard Statements':
        update_dict["hazard_statements"] = {
            'content': data_dict["hazard_statements"]["content"],
            'position': data_dict["hazard_statements"]["position"],
            'to_excel': field_response.replace('*','').replace('#','')
        }

    print("Base field response:", field_response, "\n")

    # Iterate over each pictogram/hazard field and mark with 'X' if applicable
    for field in fields_list:
        request_images = f"""
        Answer the question based only on this information: {field_response}.
        In the provided context, is there a {data_dict[field]["content"]} risk explicitly mentioned?
        Answer only 'X' if True, else '' (empty string).
        For the Serious health hazard risk, answer 'X' only if there is an extreme danger.
        """

        to_excel_value = model.predict(request_images)
        print(data_dict[field]["content"], to_excel_value)

        # Update dictionary
        update_dict[field] = {
            'content': data_dict[field]["content"],
            'position': data_dict[field]["position"],
            'to_excel': to_excel_value
        }

    # Finally, write the responses into the original dictionary
    for field, value in update_dict.items():
        data_dict[field]['to_excel'] = value['to_excel']


# Storage Fields
def candidate_lines(text: str):
    """
    Cleans and extracts candidate lines from a block of text.
    This function:
    - Removes leading/trailing whitespace.
    - Strips common bullet or list symbols (e.g., '-', '*', '•').
    - Ignores empty lines.
    - Returns a list of cleaned lines suitable for further processing,
      such as pattern matching or data extraction.
    Args:
        text (str): The raw text block to process.
    Returns:
        list[str]: A list of cleaned, non-empty lines.
    """

    lines = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        line = re.sub(r"^[\-\*\•\u2022]+\s*", "", line).strip()
        lines.append(line)
    return lines

def matches_any(line: str, patterns):
    """
    Determines whether a given line of text matches any of the provided regex patterns.
    Args:
        line (str): A single line of text to be checked.
        patterns (list[str]): A list of regular expression patterns to match against.
    Returns:
        bool: True if the line matches at least one pattern, False otherwise.
    Notes:
        - The match is case-insensitive since the line is converted to lowercase.
        - Useful for filtering or identifying relevant lines in text extraction tasks.
    """

    line_lower = line.lower()
    return any(re.search(p, line_lower) for p in patterns)

def find_support_for_storage(base_summary: str, field_key: str) -> str:
    """
    Finds supporting evidence for a specific storage-related field in a summary text.
    Args:
        base_summary (str): The text summary to search within.
        field_key (str): The storage field key to find evidence for (used to look up regex patterns).
    Returns:
        str: The first line from the summary that matches any pattern associated with the field.
             Returns an empty string if no match is found.
    Notes:
        - Uses _STORAGE_PATTERNS (a dictionary mapping field keys to regex patterns) for matching.
        - Lines are preprocessed with candidate_lines to remove bullets, whitespace, and list symbols.
        - Matches are case-insensitive via matches_any.
    """

    patterns = _STORAGE_PATTERNS.get(field_key, [])
    if not patterns:
        return ""
    for line in candidate_lines(base_summary):
        if matches_any(line, patterns):
            return line.strip()
    return ""

def extract_json_block(text: str):
    """
    Extracts a JSON block from a given text string.
    Args:
        text (str): Input text that may contain a JSON object, possibly wrapped in ```json ... ```.
    Returns:
        str: The JSON block as a string if found; otherwise, returns the original input text.
    Notes:
        - Uses a regular expression to detect the first valid JSON object at the end of the text.
        - Supports multi-line JSON by using DOTALL mode.
        - If no JSON object is detected, the function safely returns the original text.
    """

    match = re.search(r"\{.*\}\s*$", text, re.DOTALL)
    if match:
        return match.group(0)
    return text

def storage_fields_with_images(field_name, content, fields_list, data_dict, use_llm=True, model=None):
    print("Entro a Storage")
    """
    Processes storage-related fields in an SDS/MSDS context.
    - Step 1 (PPE-Storage): Marks 'X' in `to_excel` for 7 predefined STORAGE fields.
      Stores the supporting line from the base summary in `response`.

    - Step 2 (Other storage measures): Handles `special_storage_describe`.
      Returns only storage measures that are not part of the 7 predefined fields, in JSON:
        { "list": [...], "paragraph": "..." }
      - `to_excel` = '; '.join(list)
      - `response` = paragraph (or joined list if paragraph is empty)
    Args:
        field_name (str): Name of the storage section (e.g., "Storage").
        content (str): Full extracted text from the document.
        fields_list (list): List of JSON keys corresponding to storage fields.
        data_dict (dict): JSON structure to update with `to_excel` and `response`.
        llm: LLM object with a .invoke() method to query content.
    Returns:
        dict: Updated `data_dict` with populated storage fields.
    """
    if model is None and use_llm:
        model = default_llm

    if not isinstance(data_dict, dict):
        raise ValueError("storage_fields_with_images espera un dict (p.ej. updated_json_storage['Sheet_2']).")

        # Base summary (LLM)
    base_prompt = """
        Answer STRICTLY using only the content retrieved from the provided context.
        Do not invent or add external information.
        If the context contains no information relevant to the question, state explicitly that the information is not available.
        """
    request = f"""
        Answer the question based only on these instructions: {base_prompt}.
        What are the main {field_name} storage requirements or recommendations in the context: {content}?
        Answer in bullet points, keeping the exact wording from the context whenever possible.
        """
    try:
        base_response = model.invoke(request).content
    except Exception as e:
        print(f"storage_fields_with_images: fallo al llamar LLM para base_response: {e}")
        base_response = ""

    # Initialize fields in data_dict if missing
    for field in fields_list:
        if field not in data_dict or not isinstance(data_dict[field], dict):
            data_dict[field] = {"content": field, "position": "", "response": "", "to_excel": ""}
        else:
            data_dict[field].setdefault("content", field)
            data_dict[field].setdefault("position", "")
            data_dict[field].setdefault("response", "")
            data_dict[field].setdefault("to_excel", "")

    # Mapping of the 7 fields by LLM
    mapping_prompt = f"""
        Based only on this extracted information:
        {base_response}

        Check which of the following STORAGE requirements are explicitly required or implied.
        Mark with 'X' if true, otherwise '' (empty string).

        Fields:
        - flammables_cupboard: store in flammables cabinet/cupboard; keep away from ignition sources/heat
        - corrosives_cupboard: store in corrosives cabinet/cupboard; acids/bases segregation
        - poisons_cupboard: store in poisons/toxics cabinet; locked storage
        - ventilated_storage: ventilated storage, well-ventilated place, fume hood area
        - gas_cylinder: gas cylinders handling/storage, upright, secured, caps on
        - cold_storage: ONLY IF refrigeration or cold room is explicitly stated (e.g., "refrigerate", "cold storage", "store at/below ≤10°C", "2–8°C").
                    NOT phrases like "keep cool", "store in a cool, dry/well-ventilated place".
        - dessicated_storage: desiccator, dry storage, keep dry, protect from moisture

        Respond EXACTLY with lines like:
        field_name: X
        field_name:
        (one per line; no extra commentary)
        """
    try:
        result = model.invoke(mapping_prompt).content
    except Exception as e:
        print(f"storage_fields_with_images: fallo al llamar LLM para mapping_prompt: {e}")
        result = ""

    # Initialize as empty before marking
    for field in fields_list:
        data_dict[field]["to_excel"] = data_dict[field].get("to_excel", "")
        data_dict[field]["response"] = data_dict[field].get("response", "")

    # Mark 'X' and add evidence
    for line in result.splitlines():
        if ":" not in line:
            continue
        field, value = line.split(":", 1)
        field = field.strip().lstrip("-").strip()
        value = value.strip()
        if field in fields_list and field != "special_storage_describe":
            if value == "X":
                data_dict[field]["to_excel"] = "X"
                try:
                    data_dict[field]["response"] = find_support_for_storage(base_response, field) or ""
                except Exception:
                    data_dict[field]["response"] = ""

    # Prompt "special_storage_describe" (other measures)
    active_fields = [f for f in STORAGE_FIELDS if data_dict.get(f, {}).get("to_excel") == "X"]
    other_prompt = f"""
        You are given this extracted text (context):
        {base_response}

        Task: List ALL explicit storage measures that are NOT any of these categories:
        - flammables cupboard (flammables cabinet/cupboard; ignition sources)
        - corrosives cupboard (corrosives cabinet; acids/bases segregation)
        - poisons cupboard (toxics cabinet; locked storage)
        - ventilated storage (well-ventilated place; fume hood area)
        - gas cylinder storage (upright; secured; caps on)
        - cold storage (refrigerate; keep cool; temp control)
        - dessicated storage (desiccator; keep dry; protect from moisture)

        Also, avoid rephrasing measures already covered by active categories: {active_fields}.

        Return STRICT JSON with this schema:
        {{
          "list": ["short, practical storage measure 1", "short, practical storage measure 2", ...],
          "paragraph": "exact paragraph(s) from the context where those 'other' storage measures appear"
        }}

        Requirements:
        - "list": concise, actionable phrases; do not include H-codes or P-codes; no generic headings.
        - "paragraph": copy the original text fragment(s) verbatim from the context containing those 'other' storage measures.
        - If there are none, return: {{ "list": [], "paragraph": "" }}.
        """
    try:
        other_raw = model.invoke(other_prompt).content
    except Exception as e:
        print(f"storage_fields_with_images: fallo al llamar LLM para other_prompt: {e}")
        other_raw = ""

    # Robust JSON parsing
    other_json = {"list": [], "paragraph": ""}
    try:
        block = extract_json_block(other_raw)
        parsed = json.loads(block)
        if isinstance(parsed, dict):
            other_json = parsed
    except Exception:
        other_json = {"list": [], "paragraph": ""}

    measures_list = other_json.get("list", [])
    if not isinstance(measures_list, list):
        measures_list = []
    paragraph = other_json.get("paragraph", "")
    if not isinstance(paragraph, str):
        paragraph = ""

    # Write special_storage_describe (if it exists)
    if "special_storage_describe" in fields_list:
        to_excel_val = "; ".join([m.strip() for m in measures_list if isinstance(m, str) and m.strip()])
        data_dict["special_storage_describe"]["to_excel"] = to_excel_val
        # response: prefer the original paragraph; if empty, use the joined list
        data_dict["special_storage_describe"]["response"] = paragraph.strip() if paragraph.strip() else to_excel_val

    # Return the updated dict (since the caller passes updated_json_storage['Sheet_2'])
    return {"Sheet_2": data_dict}


# Hazard Group
# Function to update the hazard group in a JSON structure
def update_hazard_group_in_json(data, hazard_letter):
    """
    Recursively traverses a nested JSON structure and updates all occurrences
    of the 'hazard_group' field by setting its 'response' and 'to_excel' values.
    Args:
        data (dict or list): Nested JSON structure containing hazard fields.
        hazard_letter (str): The hazard group letter (e.g., 'A', 'B', 'C') to assign.
    Behavior:
        - If a dictionary has a key 'hazard_group' and its value is a dictionary,
          sets both 'response' and 'to_excel' to `hazard_letter`.
        - Recursively processes nested dictionaries and lists.
    Returns:
        The updated JSON structure with all 'hazard_group' fields modified.
    """
    if isinstance(data, dict):
        for key, value in data.items():
            if key == "hazard_group" and isinstance(value, dict):
                value["response"] = hazard_letter
                value["to_excel"] = hazard_letter
            else:
                update_hazard_group_in_json(value, hazard_letter)
    elif isinstance(data, list):
        for item in data:
            update_hazard_group_in_json(item, hazard_letter)
    return data

def fill_hazard_group_rag(source_match, json_input, content="") -> Dict[str, Any]:
    """
    Extracts hazard H-codes from a document and assigns a hazard letter (A to N) to all 'hazard_group' fields in a JSON.
    Steps:
    1. Extract all H-codes (e.g., H300, H315) from the text `content`.
    2. Map each code to a hazard letter using a global `hazard_classification` table.
       - Fallback classification is provided if not defined globally.
       - Unknown codes default to letter 'E'.
    3. Determine the most severe hazard letter among all detected codes (priority A > B > C > D > E > N).
    4. Recursively update the JSON:
       - Each 'hazard_group' dictionary has its 'response' and 'to_excel' set to the selected letter.
    Args:
        json_input (dict): Nested JSON containing hazard fields to update.
        source_match (str): Source document identifier (used in logging).
        content (str): Text content of the document from which H-codes are extracted.
    Returns:
        dict: The updated JSON with all 'hazard_group' fields filled with the appropriate hazard letter.
    Notes:
        - If no H-codes are detected, assigns 'N' (lowest hazard).
        - Prints debug information including detected codes, mapped letters, and final letter.
    """

    # Global classification table (fallback if not defined)
    global hazard_classification
    if "hazard_classification" not in globals():
        hazard_classification = {
            "N": [],
            "E": ["H303", "H305", "H313", "H316", "H318", "H320", "H333"],
            "D": ["H302", "H312", "H332", "H315", "H319"],
            "C": ["H341", "H351", "H361", "H362", "H371", "H373", "H317", "H335", "H336"],
            "B": ["H301", "H304", "H311", "H331", "H334", "H314", "H318"],
            "A": ["H300", "H310", "H330", "H340", "H350", "H360", "H370", "H372"]
        }

    # Severity priority (A = most severe)
    priority = ["A", "B", "C", "D", "E", "N"]

    # Build code -> letter mapping
    code_to_letter = {code: letter for letter in priority for code in hazard_classification.get(letter, [])}

    # Extract all H-codes from the text (normalized)
    h_codes = re.findall(r'H\s*\d{3}', (content or "").upper())
    h_codes = [c.replace(" ", "") for c in h_codes]

    if not h_codes:
        hazard_letter = "N"
        print(f"No H-codes detected in {source_match}")
    else:
        # Map to letters, unknown codes = "E"
        detected_letters = {code_to_letter.get(c, "E") for c in h_codes}

        # Select the most severe
        hazard_letter = next((letter for letter in priority if letter in detected_letters), "N")

        print(f"H-codes detected in {source_match}: {h_codes}")
        print(f"Letters detected: {detected_letters}")
        print(f"Final hazard letter: {hazard_letter}")

    # Update JSON recursively
    def _update_hazard_group(node):
        if isinstance(node, dict):
            for key, value in node.items():
                if key == "hazard_group" and isinstance(value, dict):
                    value["response"] = hazard_letter
                    value["to_excel"] = hazard_letter
                else:
                    _update_hazard_group(value)
        elif isinstance(node, list):
            for item in node:
                _update_hazard_group(item)

    _update_hazard_group(json_input)
    return json_input


# Severity and Probability
def fill_json_severity_probability(json_input: Dict[str, Any]) -> Dict[str, Any]:
    """
    Populates the severity and probability fields in a nested JSON for risk assessment.
    Specifically updates the following fields if present in 'Sheet_2':
    - 'severity'
    - 'likelihood_before_control_measures'
    - 'likelihood_after_control_measures'
    Args:
        json_input (dict): Nested JSON expected to contain a 'Sheet_2' key with relevant fields.
    Returns:
        dict: The same JSON with updated 'response' and 'to_excel' values for the specified fields.
    Notes:
        - Default values applied:
            • severity -> "Severe"
            • likelihood_before_control_measures -> "Possible"
            • likelihood_after_control_measures -> "Unlikely"
        - Other fields in the JSON remain untouched.
    """

    # Try Sheet_2 first, fallback to root
    sheet = json_input.get("Sheet_2", json_input)

    if not isinstance(sheet, dict):
        raise ValueError("No valid sheet found in the JSON")

    # Default values
    if isinstance(sheet.get("severity"), dict):
        sheet["severity"]["response"] = "Severe"
        sheet["severity"]["to_excel"] = "Severe"

    if isinstance(sheet.get("likelihood_before_control_measures"), dict):
        sheet["likelihood_before_control_measures"]["response"] = "Possible"
        sheet["likelihood_before_control_measures"]["to_excel"] = "Possible"

    if isinstance(sheet.get("likelihood_after_control_measures"), dict):
        sheet["likelihood_after_control_measures"]["response"] = "Unlikely"
        sheet["likelihood_after_control_measures"]["to_excel"] = "Unlikely"

    return json_input

def extract_hazards_text(source_match, json_input, use_llm=True, model=None, content="", fields_list=None) -> Dict[str, Any]:
    """
    Extracts hazard-related information from an SDS/MSDS document and populates a JSON.
    This function focuses on extracting the following hazard fields:
    - physical_form_and_quantity
    - potential_routes_of_exposure
    - workplace_exposure_limits
    - arising_harm
    It uses hierarchical prompting with an LLM to first select the relevant context
    and then generate a full answer, including a concise Excel-ready summary.
    Args:
        source_match (str): Filename or document reference for tracking/logging.
        json_input (dict): JSON structure to populate (must contain 'Sheet_2').
        llm: LLM object with a .predict() method to query.
        content (str): Full SDS/MSDS document text.
        fields_list (List[str], optional): Specific fields to extract. Defaults to all hazard fields.
        table_index (int, optional): Index of the table (reserved for future use).
    Returns:
        dict: Updated JSON with 'response' (full text) and 'to_excel' (concise summary) for each field.
    Notes:
        - If no relevant information is found, 'to_excel' will be "N/A".
        - EXCEL_SUMMARY is generated in English, max 50 words / 200 characters.
        - The LLM is strictly instructed not to invent any information.
    """

    excel_na_to_excel = "N/A"
    sheet_key = "Sheet_2"
    max_excel_chars = 300

    if model is None and use_llm:
        model = default_llm

    if sheet_key not in json_input:
        raise ValueError(f"Sheet key '{sheet_key}' not found in json_input")

    # Questions for each field
    questions = {
        "physical_form_and_quantity": (
            "What is the physical form of the substance (gas, liquid, solid) "
            "and in what packaging or quantity format is it supplied (e.g., bottle 200 ml, bag, sack, cylinder)?"
        ),
        "potential_routes_of_exposure": (
            "What are the possible routes of exposure to the substance for humans? "
            "(e.g., inhalation, skin contact, eye contact, ingestion)."
        ),
        "workplace_exposure_limits": (
            "What are the Workplace Exposure Limits (WEL), TWA (8h), STEL (15 min), or other exposure thresholds "
            "provided? Include numeric values and units."
        ),
        "arising_harm": (
            "What are the potential harms or adverse effects associated with exposure to this substance? "
            "(e.g., toxic effects, respiratory issues, organ damage, skin/eye irritation)."
        )
    }

    # Context selector prompt
    prompt_selector = (
        "You are an intelligent assistant specialized in analyzing Safety Data Sheets (SDS).\n"
        "You will be given the full SDS text and a SECTION name.\n"
        "Task: Extract only the sentences and fragments relevant to the SECTION. "
        "If nothing is relevant, return an empty string.\n\n"
        "SECTION: {section}\n\n"
        "DOCUMENT:\n{fragments}\n\n"
        "Return ONLY the relevant CONTEXT text."
    )

    # Final response prompt template
    prompt_template = (
        "You are a precise assistant specialized in Safety Data Sheets.\n"
        "Answer STRICTLY using only the provided CONTEXT. Do not invent or add external info.\n\n"
        "If the document contains no relevant information, answer: 'The document does not provide this information.'\n\n"
        "At the end of your answer, ALWAYS add a final line:\n"
        "EXCEL_SUMMARY: <short summary or 'no information'>\n\n"
        "Rules for EXCEL_SUMMARY:\n"
        " - If no info: EXCEL_SUMMARY: no information\n"
        " - If info exists: concise summary (max 50 words, max 200 chars), in English\n"
        " - Prefer keywords, numbers, hazard codes, short phrases, comma-separated\n\n"
        "QUESTION: {question}\n\n"
        "CONTEXT:\n{context}"
    )

    excel_marker_re = re.compile(r"EXCEL_SUMMARY:\s*(.+)$", re.IGNORECASE | re.MULTILINE)

    # Process each field
    for field in fields_list:
        cell = json_input[sheet_key].get(field, {})
        question = questions.get(field, f"Extract the information about {field}.")

        # Step 1: Context selector
        selector_prompt = prompt_selector.format(
            section=field,
            fragments=content
        )
        try:
            context_filtered = model.predict(selector_prompt).strip()
        except Exception as e:
            context_filtered = ""
            print(f"Error in context selector for '{field}': {e}")

        # Step 2: Final answer
        final_prompt = prompt_template.format(
            question=question,
            context=context_filtered
        )
        try:
            full_response = model.predict(final_prompt).strip()
        except Exception as e:
            full_response = ""
            print(f"Error in final response for '{field}': {e}")

        # Extract EXCEL_SUMMARY
        m = excel_marker_re.search(full_response)
        if m:
            excel_summary = m.group(1).strip()
            to_excel_value = (
                excel_na_to_excel
                if excel_summary.lower() in ("no information", "not available", "no information available")
                else excel_summary[:max_excel_chars].rstrip()
            )
        else:
            to_excel_value = excel_na_to_excel

        # Update JSON
        json_input[sheet_key][field]["response"] = full_response
        json_input[sheet_key][field]["to_excel"] = to_excel_value

        print("-----"*80)
        print(f"Field: '{field}'")
        print(f"response: {full_response}")
        print(f"to_excel: {to_excel_value}")

    return json_input


def general_text_extraction (source_match, json_input, use_llm=True, model=None, content="", fields_list=None, table_index=0) -> Dict[str, Any]:
    """
    Performs hierarchical extraction of information from a full SDS/MSDS document.
    This function iterates over a list of fields in the JSON, extracts only the relevant context
    from the document for each field using a two-step LLM process (context selection + detailed answer),
    and produces an Excel-ready summary (EXCEL_SUMMARY) for each field.
    Args:
        source_match (str): Document filename or identifier for tracking/logging.
        json_input (Dict[str, Any]): JSON structure to populate (must contain 'Sheet_2').
        llm: LLM object with a .predict() method.
        content (str): Full text content of the SDS/MSDS document.
        fields_list (List[str], optional): List of field keys to extract. Defaults to all fields in 'Sheet_2'.
        table_index (int, optional): Index of the table or section for context (default is 0).
    Returns:
        Dict[str, Any]: Updated JSON with the following for each field:
            - 'response': full LLM answer for the field.
            - 'to_excel': concise, Excel-friendly summary.
    Notes:
        - EXCEL_SUMMARY is generated in English, max 50 words / 200 characters.
        - If no information is found for a field, 'to_excel' will be "N/A".
        - The function ensures the original JSON structure is preserved.
        - Strictly instructs the LLM not to invent information beyond the document content.
    """

    if model is None and use_llm:
        model = default_llm

    excel_na_to_excel = "N/A"
    sheet_key = "Sheet_2"
    max_excel_chars = 300

    if fields_list is None:
        fields_list = list(json_input[sheet_key].keys())

    # Use full document content
    full_document_text = content.strip()

    # Prompts
    prompt_selector = (
        "You are an intelligent assistant specialized in analyzing Safety Data Sheets (SDS).\n"
        "You will be given a full SDS document and a target SECTION name.\n"
        "Task: Read the document, understand the whole context, and produce a single coherent CONTEXT text\n"
        "that contains only the information relevant to the SECTION. If there is no relevant information, return an empty string.\n\n"
        "SECTION: {section}\n\n"
        "DOCUMENT:\n{fragments}\n\n"
        "Return ONLY the CONTEXT text (no JSON, no explanation)."
    )

    prompt_template = (
        "You are a precise technical assistant specialized in Safety Data Sheets. "
        "Answer STRICTLY using only the content retrieved from the provided document. "
        "Do not invent or add external information. If the document contains no information "
        "relevant to the question, state explicitly that the information is not available.\n\n"
        "If the QUESTION requests only 'details' without specifying more, provide a comprehensive summary of ALL information "
        "received from the document or retrieved context.\n\n"
        "At the end of your answer, ALWAYS add a final line that starts exactly with:\n"
        "EXCEL_SUMMARY: <one-line summary or 'no information'>\n\n"
        "Rules for the EXCEL_SUMMARY (VERY IMPORTANT):\n"
        " - If there is no relevant information, write exactly: EXCEL_SUMMARY: no information\n"
        " - If there IS relevant information, use a concise summary suitable for a single Excel cell:\n"
        "   • Keep it very short: max 50 words and max 200 characters.\n"
        "   • Prefer keywords, numeric values, hazard codes (e.g., H315), or short phrases.\n"
        "   • If multiple small items, use comma-separated short phrases (no newlines).\n"
        " - The EXCEL_SUMMARY must always be in English.\n\n"
        "Now answer the QUESTION using only document content."
    )

    excel_marker_re = re.compile(r"EXCEL_SUMMARY:\s*(.+)$", re.IGNORECASE | re.MULTILINE)

    # Section name
    section_name = dtr_tables[table_index]

    # Process each field
    for campo in fields_list:
        cell = json_input[sheet_key].get(campo, {})
        consulta = str(cell.get("content", "") or "").strip()

        if not consulta:
            json_input[sheet_key][campo] = {
                "content": cell.get("content", ""),
                "position": cell.get("position", ""),
                "response": "",
                "to_excel": excel_na_to_excel
            }
            continue

        # Step 1: Context selection
        selector_prompt = prompt_selector.format(
            section=section_name,
            fragments=full_document_text
        )
        try:
            context_filtered = model.predict(selector_prompt).strip()
        except Exception as e:
            context_filtered = ""
            print(f"Error in context selector for field '{campo}': {e}")

        # Step 2: Final response
        final_prompt = (
            f"{prompt_template}\n\nQUESTION: {consulta}\n\nCONTEXT:\n{context_filtered}"
        )
        try:
            respuesta_completa = model.predict(final_prompt).strip()
        except Exception as e:
            respuesta_completa = ""
            print(f"Error in final response for field '{campo}': {e}")

        # Extract EXCEL_SUMMARY
        m = excel_marker_re.search(respuesta_completa)
        if m:
            excel_summary = m.group(1).strip()
            to_excel_value = (
                excel_na_to_excel
                if excel_summary.lower() in ("no information", "no information available", "not available")
                else excel_summary[:max_excel_chars].rstrip()
            )
        else:
            to_excel_value = excel_na_to_excel

        # Save in JSON
        json_input[sheet_key][campo]["response"] = respuesta_completa
        json_input[sheet_key][campo]["to_excel"] = to_excel_value

        # Source
        print("-----"*80)
        print(f"Field: '{campo}'")
        print(f"response: {respuesta_completa}")
        print(f"to_excel: {to_excel_value}")
    return json_input


# Function to fill an Excel with JSON data, applying Arial 12, centered text, and wrap text
def fill_excel_with_json(jsons_list: list, template_path: str, output_dir: str, source_match: str):
    """
    Fills an Excel template with data from multiple JSON objects and saves it as a new file.
    This function writes all values from JSON fields into an Excel template while applying
    formatting: Arial 12 font, bold, centered text, and wrap text. Column widths are
    automatically adjusted based on content length.
    Args:
        jsons_list (list): List of JSON objects containing fields with 'position' and 'to_excel'.
        template_path (str): Path to the Excel template file (.xlsx).
        output_dir (str): Folder where the output Excel file will be saved.
        source_match (str): Base name of the file used to generate the final output filename.
    Notes:
        - Only cell values are modified; the template's style/formatting is preserved.
        - For fields 'likelihood_before_control_measures', 'severity', and
          'likelihood_after_control_measures', 'position' can be a list and the same value
          will be written to all positions.
        - All text is written in Arial 12, bold, centered, and wrapped.
        - Column widths are adjusted automatically based on the maximum content length.
        - Output filename format: {source_match}_YYYY-MM-DD_HHMM.xlsx
    Returns:
        str: Full path of the generated Excel file.
    """

    wb = openpyxl.load_workbook(template_path)
    ws = wb['COSHH Assessment']

    standard_font = Font(name="Arial", size=12, bold=True)
    standard_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    multi_position_fields = {
        "likelihood_before_control_measures",
        "severity",
        "likelihood_after_control_measures"
    }

    for json_data in jsons_list:
        for field, values in json_data.get("Sheet_2", {}).items():
            position = values.get("position")
            to_excel = values.get("to_excel")

            if not position or not to_excel:
                continue

            if field in multi_position_fields and isinstance(position, list):
                for pos in position:
                    cell = ws[pos]
                    cell.value = to_excel
                    cell.font = standard_font
                    cell.alignment = standard_alignment
            elif isinstance(position, str):
                cell = ws[position]
                cell.value = to_excel
                cell.font = standard_font
                cell.alignment = standard_alignment

    datetime_str = datetime.now().strftime("%Y-%m-%d_%H%M")
    filename = f"{source_match}_{datetime_str}.xlsx"
    output_path = os.path.join(output_dir, filename)

    wb.save(output_path)
    wb.close()
    return str(output_path)
